"""
Genera el Excel de salida: Estado de Resultados historico vs ajustado por
inflacion, cuenta por cuenta, con totales.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NUM_FORMAT = "#,##0.00"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
RECPAM_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_workbook(df_result, meses: list, cliente: str, indices: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "EERR Ajustado"

    # --- Encabezado informativo ---
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Estado de Resultados ajustado por inflacion - {cliente}"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A2"] = "Generado el:"
    ws["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A3"] = "Ejercicio (meses detectados):"
    ws["B3"] = " -> ".join(meses)

    start_row = 5

    # --- Fila de titulos de grupo (mes) ---
    col = 3  # A=1 codigo, B=2 nombre, luego pares Historico/Ajustado por mes
    ws.cell(row=start_row, column=1, value="Codigo")
    ws.cell(row=start_row, column=2, value="Cuenta")
    ws.cell(row=start_row + 1, column=1, value="")
    ws.cell(row=start_row + 1, column=2, value="")

    for mes in meses:
        ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row, end_column=col + 1)
        ws.cell(row=start_row, column=col, value=mes)
        ws.cell(row=start_row + 1, column=col, value="Historico")
        ws.cell(row=start_row + 1, column=col + 1, value="Ajustado")
        col += 2

    ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row, end_column=col + 1)
    ws.cell(row=start_row, column=col, value="TOTAL")
    ws.cell(row=start_row + 1, column=col, value="Historico")
    ws.cell(row=start_row + 1, column=col + 1, value="Ajustado")
    total_col_start = col

    last_col = col + 1

    for c in range(1, last_col + 1):
        for r in (start_row, start_row + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True, color="FFFFFF" if r == start_row else "000000")
            cell.fill = HEADER_FILL if r == start_row else SUBHEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

    # --- Fila de indices utilizados ---
    idx_row = start_row + 2
    ws.cell(row=idx_row, column=2, value="Indice FACPCE utilizado")
    ws.cell(row=idx_row, column=2).font = Font(italic=True, size=9)
    col = 3
    for mes in meses:
        ws.cell(row=idx_row, column=col, value=indices.get(mes))
        ws.merge_cells(start_row=idx_row, start_column=col, end_row=idx_row, end_column=col + 1)
        c = ws.cell(row=idx_row, column=col)
        c.font = Font(italic=True, size=9, color="666666")
        c.alignment = Alignment(horizontal="center")
        col += 2

    # --- Filas de datos ---
    data_start = idx_row + 1
    r = data_start
    for _, fila in df_result.iterrows():
        ws.cell(row=r, column=1, value=fila["codigo"])
        ws.cell(row=r, column=2, value=fila["nombre"])
        col = 3
        for mes in meses:
            ws.cell(row=r, column=col, value=round(float(fila[mes]), 2)).number_format = NUM_FORMAT
            ws.cell(row=r, column=col + 1, value=round(float(fila[f"{mes}_ajustado"]), 2)).number_format = NUM_FORMAT
            col += 2
        ws.cell(row=r, column=total_col_start, value=round(float(fila["total"]), 2)).number_format = NUM_FORMAT
        ws.cell(row=r, column=total_col_start + 1, value=round(float(fila["total_ajustado"]), 2)).number_format = NUM_FORMAT
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    data_end = r - 1

    # --- Fila de totales ---
    ws.cell(row=r, column=2, value="TOTAL")
    ws.cell(row=r, column=2).font = Font(bold=True)
    col = 3
    for mes in meses:
        for offset in (0, 1):
            col_letter = get_column_letter(col + offset)
            cell = ws.cell(row=r, column=col + offset)
            cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            cell.number_format = NUM_FORMAT
            cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL
        col += 2
    for offset in (0, 1):
        col_letter = get_column_letter(total_col_start + offset)
        cell = ws.cell(row=r, column=total_col_start + offset)
        cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        cell.number_format = NUM_FORMAT
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
    for c in range(1, last_col + 1):
        ws.cell(row=r, column=c).border = BORDER

    total_row = r

    # --- Fila RECPAM (diferencia entre total ajustado y total historico) ---
    recpam_row = total_row + 1
    hist_letter = get_column_letter(total_col_start)
    aj_letter = get_column_letter(total_col_start + 1)

    ws.merge_cells(start_row=recpam_row, start_column=2, end_row=recpam_row, end_column=total_col_start - 1)
    ws.cell(row=recpam_row, column=2, value="Total RECPAM de Resultado del Ej.")
    ws.cell(row=recpam_row, column=2).font = Font(bold=True)
    ws.cell(row=recpam_row, column=2).alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(start_row=recpam_row, start_column=total_col_start, end_row=recpam_row, end_column=total_col_start + 1)
    recpam_cell = ws.cell(row=recpam_row, column=total_col_start)
    recpam_cell.value = f"={aj_letter}{total_row}-{hist_letter}{total_row}"
    recpam_cell.number_format = NUM_FORMAT
    recpam_cell.font = Font(bold=True)

    for c in range(2, last_col + 1):
        ws.cell(row=recpam_row, column=c).fill = RECPAM_FILL
        ws.cell(row=recpam_row, column=c).border = BORDER

    # --- Anchos de columna ---
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 38
    for c in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.freeze_panes = ws.cell(row=data_start, column=3)

    return wb


def export_to_bytes(df_result, meses: list, cliente: str, indices: dict) -> bytes:
    wb = build_workbook(df_result, meses, cliente, indices)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
